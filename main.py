from selenium import webdriver
from openpyxl import load_workbook
from modelo.login import Usuario
from modelo.navegar import Navegacao
from modelo.scraping import Scrapy

path_file = "\\downloads"

prefs = {"safebrowsing.enabled": True}
prefs2 = {"download.default_directory": path_file}

prefs.update(prefs2)

options = webdriver.ChromeOptions()

options.add_argument("-incognito")
options.add_argument("--disable-dev-shm-usage")
options.add_argument("start-maximized")
options.add_argument("--no-sandbox")
options.add_experimental_option("prefs", prefs)
bot = webdriver.Chrome(options=options, port=5556)

bot.get("http://emec.mec.gov.br/modulos/visao_comum/php/login/comum_login.php?")

caminho_login = './dados/login_e_senha.xlsx' #<------------------------------------------------- Altera o Caminho Login
excel_login = load_workbook(caminho_login)
login_plan = excel_login.active
usu = login_plan.cell(row=1, column=2).value
password = login_plan.cell(row=2, column=2).value

usuario = Usuario(bot)

usuario.clicando_na_ies()
usuario.username(usu)
usuario.senha(password)
usuario.entrar_no_sistema()

navegar = Navegacao(bot)

navegar.clicando_na_unesa()

bot.get("http://emec.mec.gov.br/emec/ies/cadastro-institucional/index")


caminho_cod = './dados/input.xlsx' #<------------------------------------------------- Altera o Caminho Input
excel_cod = load_workbook(caminho_cod)
cod_plan = excel_cod.active
qtd_line_cod = cod_plan.max_row

scraping = Scrapy(bot)

for i in range(2, qtd_line_cod + 1):
    cod = cod_plan.cell(row=i, column=1).value
    curso = cod_plan.cell(row=i, column=2).value
    qtd_curso = cod_plan.cell(row=i, column=3).value
    if i == 2:
        navegar.validar_loading()

    navegar.filtrar_por_codigo()
    navegar.pesquisar_por_codigo(cod)
    navegar.clicar_no_curso(cod)

    scraping.selecionar_max_page(qtd_curso)

    scraping.tabela(cod, curso)
    scraping.fechar_tabela()


