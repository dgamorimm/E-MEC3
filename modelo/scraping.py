from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support.expected_conditions import presence_of_element_located, \
visibility_of_element_located, presence_of_element_located_s
from time import sleep
from selenium.webdriver.support.ui import Select
from openpyxl import load_workbook
import pandas as pd

class Scrapy:
    def __init__(self, driver):
        self.driver = driver
        self.SELECT_PAGINACAO= (By.ID, "paginationItemCountItemlista-polos")
        self.TABELA = (By.CLASS_NAME, "centralizado")
        self.ID_DIN_DATA = ""
        self.ID_DIN_VAGAS = ""
        self.ClICK_FECHAR = (By.ID, "btnDialog_0")

        self.caminho_out = './dados/output.xlsx'  # <------------------------------------------------- Altera o Caminho Input
        self.excel_out = load_workbook(self.caminho_out)
        self.out_plan = self.excel_out.active
        self.qtd_line_out = self.out_plan.max_row

    def find(self, *locator, timeout=30):
        return WebDriverWait(self.driver, timeout).until(presence_of_element_located(*locator))

    def finds(self, *locator, timeout=30):
        return WebDriverWait(self.driver, timeout).until(presence_of_element_located_s(*locator))

    def find_reduce(self, *locator, timeout=3):
        return WebDriverWait(self.driver, timeout).until(presence_of_element_located(*locator))

    def find_v(self, *locator, timeout=2):
        return WebDriverWait(self.driver, timeout).until(visibility_of_element_located(*locator))


    def frame_switch_id(self, id):
        driver = self.driver
        driver.switch_to.frame(driver.find_element_by_id(id))

    def frame_switch_name(self, name):
        driver = self.driver
        driver.switch_to.frame(driver.find_element_by_name(name))

    def validar_click(self,clique_certo = None):
        try:
            self.find(clique_certo).click()
        except Exception as x:
            print(x)
            sleep(2)
            return self.validar_click(clique_certo)

    def validar_click_reduce(self,clique_certo = None):
        try:
            self.find_reduce(clique_certo).click()
        except IOError as x:
            print(x)
            sleep(2)
            return self.validar_click_reduce(clique_certo)

    def validar_click_imediato(self,clique_certo = None):
        try:
            clique_certo.click()
        except IOError as x:
            print(x)
            sleep(2)
            return self.validar_click_imediato(clique_certo)

    # ajustar a paginação
    def selecionar_max_page(self, count_curso):
        try:
           select_page = Select(self.find_reduce(self.SELECT_PAGINACAO))
           if int(count_curso) > 30:
            select_page.select_by_value("/emec/ies/cadastro-institucional/listar-polos/list/1000")
            return sleep(2)
        except:
            return sleep(2)

    def tabela(self, cod_curso, curso):
        tabela = self.finds(self.TABELA)
        excel = pd.read_excel("dados/output.xlsx", encoding="latin-1")
        i = excel.shape[0] + 2
        for item in tabela:
            if (str(item.text) not in ["Código","Polo",""]):
                print(cod_curso)
                self.out_plan.cell(row= i, column=1).value = cod_curso

                print(curso)
                self.out_plan.cell(row= i, column= 2).value = curso

                cod = item.text
                print(cod)
                self.out_plan.cell(row= i, column= 3).value  = cod

                self.ID_DIN_DATA = (By.ID, "dt_funcionamento_" + cod)
                data = self.find(self.ID_DIN_DATA)
                data_inicio = data.get_attribute("value")
                print(data_inicio)
                self.out_plan.cell(row= i, column= 4).value = data_inicio

                self.ID_DIN_VAGAS = (By.ID, "nu_vagas_" + cod)
                vg = self.find(self.ID_DIN_VAGAS)
                vaga = vg.get_attribute("value")
                print(vaga)
                self.out_plan.cell(row=i, column=5).value = vaga

                i += 1
                self.excel_out.save("./dados/output.xlsx")
        pass

    def fechar_tabela(self):
        close = self.finds(self.ClICK_FECHAR)
        for fch in close:
            if fch.text == 'FECHAR':
                fch.click()
        return sleep(1)

